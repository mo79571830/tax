import openpyxl
import requests
from bs4 import BeautifulSoup

url = "http://www.cq-l-tax.gov.cn/MiddleQuery/nsxypg_jg_gs.jsp"


def get_tax_num():
    taxes = []
    __file = '/Users/liuyihong/Desktop/p.txt'
    # __file = 'C:/p.txt'
    with open(__file) as f:
        for line in f.readlines():
            taxes.append(line.strip())
    return taxes


def get_data(num):
    __parma = {'qNsrsbh': num,
               'qNd': '2017',
               'act': 'search'}

    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,mage/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "close",
        "Content-type": "application/x-www-form-urlencoded",
        "Host": "www.cq-l-tax.gov.cn",
        "Referer": "http://www.cq-l-tax.gov.cn/MiddleQuery/nsxypg_jg_gs.jsp",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) \
                Chrome/70.0.3538.67 Safari/537.36"
    }

    r = requests.post(url, __parma, headers=headers)
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'html.parser')
    values = soup.find_all('table')[1].find_all('label')
    value = BeautifulSoup(str(values), 'html.parser').text.strip().strip('[]').replace("\r\n", "").split(",")
    print(value)
    result = value[1::2]
    return result


def write_into_execl(value, index):
    __file = '/Users/liuyihong/Desktop/test.xlsx'
    # __file = 'C:/test.xlsx'
    wb = openpyxl.load_workbook(__file, read_only=False)
    wb.create_sheet('test')
    sheet = wb.get_sheet_by_name('工作表1')
    # sheet.cell(row=row, column=cow, value=value)
    sheet[index] = value


if __name__ == '__main__':
    tax_nums = get_tax_num()
    f = open('/Users/liuyihong/Desktop/123.txt', 'a')
    for val in tax_nums:
        datas = get_data(val)
        string = ','.join(datas) + '\n'
        f.writelines(string)

    f.close()

    # rows = 1
    # cows = 97
    # index = ''
    # for data in datas:
    #     index = chr(cows) + str(rows)
    #     write_into_execl(data, index)
    #     cows += 1
    # rows += 1
